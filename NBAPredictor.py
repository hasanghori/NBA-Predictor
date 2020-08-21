from openpyxl import load_workbook

# Notes to Keep in Mind
# Below is the order in which the game stats are organized
# ["0:Team", 1:Matchup (featuring team1, team2), 2:Date, 3:Win/Loss, 4:mins, 5:"points", 6:"FGM", 7:"FGA", 8:"FG%", 9:"3PM", 10:"3PA", 11:"3P%", 12:"FTM", 13:"FTA", 14:"OREB", 15:"DREB", 16:"TREB", 17:"AST", 18:"STL", 19:"BLK", 20:"TO", 21:"PF"]
# The main numbers that will pop up are
# 0: team
# [1][with either 0 or 2 inside] symbolizes the matchup. O is the main team, 1 is the opponent
# [5] that is where all the main stats begin from

"""I want to give a quick overview of how this program works and what it does
    The main purpose of this program is to see how accurately I could predict the results of the NBA bubble games
    To begin, the program will prompt you to insert the number of games you wish to analyze. The program will then continuesly calculate the average of a team over those last X games.
    This simulates the bubbles 8 game regular season games and will allow for accurate predictions over the playoffs
    After inputting the number of games, the program will calculate the percent of games you could predict correctly if you just decided, "I will choose the team that averages the most in X stat"
    This creates a baseline percentage. My goal is to A) beat predictions made by sports betting sites and B) to beat predictions someone could make by just choosing the team that averages more points or something like that
    This program looks at two stats and calculates the probability of a team winning based on having two stat averages above others then sorts these probabilities in a list based on the highest probability to the lowest
    For example, lets say you inputted 8 for your "number of games you wish to analyze". This program will scroll through all the NBA games in the last 2 years and calculate the amount of times the team that averaged (for example) more points and assists won over the last 8 games, won the game. 
    It will conduct this analysis for every game over the last 2 years and for every combination of two stats together
    Finally when you when you input two teams that you wish to calculate the probability of winning, the program will look at both teams and their averages over the last X games.
    Then scroll through the list of available probabilities and find if one of teams is averaging a higher amount in two stats that have a specified probability in the list available
    The program will then return the team that has the better averages and the corresponding probability
     
    Note: to break even in sports betting, a better must get 53% of bets right
    The best professional sports betters in the world get around 54% to 55% of games right
    This program gets 59.7% of games correct
        Note: for the bubble testing I have done it is closer to the 64% range over a 14 playoff game sample size
        Predictions will be posted on twitter @HGPredictions
        
"""


def mostImportantStatRunAvg(gamesArray, numGames):  # This element calculates the
    percentWinStat = ["Team", "", "", "", "", int(0), int(0), int(0), int(0), int(0), int(0), int(0), int(0), int(0),
                      int(0), int(0), int(0), int(0), int(0), int(0), int(0), int(
            0)]  # this is an array of all the stats so we can see the individual percentages of choosing who avgs the most for each one
    totalGame = int(0)
    for team in range(0,
                      30):  # the array that stores games is a 2d list that is sorted in 1 dimension by team (30 teams in NBA) and the other Direction by amount of games played (154)
        for game in range(0,
                          154):  # note not every team has played 154 games so there are some blank spaces within the list
            if gamesArray[team][game][
                0] != "":  # accounts for blank spaces. If 'game' element is blank then this will not run
                avgTeam1 = AvgSinceWhen(gamesArray, gamesArray[team][game][2], gamesArray[team][game][0],
                                        numGames)  # takes average from specified date
                avgTeam2 = AvgSinceWhen(gamesArray, gamesArray[team][game][2], gamesArray[team][game][1][2],
                                        numGames)  # sorted (All Games, date, team to take avg of, numGames)
                totalGame = totalGame + 1
                for compareStat in range(5, 23):
                    if avgTeam1[compareStat] > avgTeam2[compareStat] and gamesArray[team][game][3] == 'W':
                        percentWinStat[compareStat] = percentWinStat[compareStat] + 1
                    elif avgTeam1[compareStat] < avgTeam2[compareStat] and gamesArray[team][game][3] == 'L':
                        percentWinStat[compareStat] = percentWinStat[compareStat] + 1
    for x in range(5, 22):
        percentWinStat[x] = round(int(percentWinStat[x]) / int(totalGame), 3)
    return percentWinStat


def AvgSinceWhen(gamesArray, date, team, numGames):  # takes average over the last 'numGames' from a specific date
    avgGame = ["", "", "", int(0), int(0), int(0), int(0), int(0), int(0), int(0), int(0), int(0), int(0), int(0),
               int(0), int(0), int(0), int(0), int(0), int(0), int(0), int(0), int(0), int(0)]
    statOrder = ["Team", "", "", "", "", "points", "FGM", "FGA", "FG%", "3PM", "3PA", "3P%", "FTM", "FTA", "OREB",
                 "DREB", "TREB", "AST", "STL", "BLK", "TO", "PF"]
    totalGames = int(0)
    for x in range(0, 30):
        if gamesArray[x][0][0] == team:
            for y in range(0, 154):
                if gamesArray[x][y][2] == date and gamesArray[x][y + 1][0] != "":
                    z = int(y + 1)
                    while z < (y + numGames) and z < 153:
                        if gamesArray[x][z][0] != "":
                            totalGames = totalGames + 1
                            for a in range(5, 24):
                                avgGame[a] = avgGame[a] + gamesArray[x][z][a]
                        if z < 153:
                            z = z + 1
                elif gamesArray[x][y][2] == date and gamesArray[x][y + 1][0] == "":
                    totalGames = totalGames + 1
    if totalGames == 0:
        totalGames = 1
    for stat in range(5, 24):
        avgGame[stat] = round(avgGame[stat] / totalGames, 3)
    return avgGame


def comboStatsRevized(gamesArray, gameStatOrder,
                      numGames):  # Finds the probability of each combination of stats and returns them in a sorted list
    statOne = int(5)
    statTwo = int(6)
    probStat = int(0)
    mixedValues = int(0)
    historicalMixedValues = int(40000)
    bestComboRanked = []
    number = int(0)
    currentProbStat = int(0)
    bestStat = int(0)
    while statOne < 21:
        statTwo = statOne + 1
        while statTwo < 22:
            imRight = int(0)
            total = int(0)
            wrong = int(0)
            for team in range(0, 30):
                for game in range(0, 154):
                    if gamesArray[team][game][0] != "":
                        # avgTeam1 = avgLastXGames(gamesArray, gamesArray[team][game][0], numGames, gamesArray[team][game][2])
                        # avgTeam2 = avgLastXGames(gamesArray, gamesArray[team][game][1][2], numGames, gamesArray[team][game][2])
                        avgTeam1 = AvgSinceWhen(gamesArray, gamesArray[team][game][2], gamesArray[team][game][0],
                                                numGames)
                        avgTeam2 = AvgSinceWhen(gamesArray, gamesArray[team][game][2], gamesArray[team][game][1][2],
                                                numGames)
                        if avgTeam1[statOne] > avgTeam2[statOne] and avgTeam1[statTwo] > avgTeam2[statTwo] and \
                                gamesArray[team][game][3] == 'W':
                            imRight = imRight + 1
                        elif avgTeam1[statOne] < avgTeam2[statOne] and avgTeam1[statTwo] < avgTeam2[statTwo] and \
                                gamesArray[team][game][3] == 'W':
                            wrong = wrong + 1
                        elif avgTeam1[statOne] < avgTeam2[statOne] and avgTeam1[statTwo] < avgTeam2[statTwo] and \
                                gamesArray[team][game][3] == 'L':
                            imRight = imRight + 1
                        elif avgTeam1[statOne] > avgTeam2[statOne] and avgTeam1[statTwo] > avgTeam2[statTwo] and \
                                gamesArray[team][game][3] == 'L':
                            wrong = wrong + 1

            # occurences = (total) - (imRight + wrong)
            if imRight > 0:
                occurences = int(imRight + wrong)
                statsAndProb = []
                currentProbStat = imRight / (imRight + wrong)
            if len(bestComboRanked) >= 1 and len(
                    bestComboRanked) < 127:  # 127 number chosen by me. It is the amount of probabilities that will be listed. I found that after 127, the probabilities were greater for the opponents to win so I wanted to leave those percentages out
                # print(number)
                run = False
                statsAndProb = [statOne, statTwo, currentProbStat,
                                occurences]  # occurences placed as a way to see if the data is reliable. the fewer occurences of a specified stat combo decreases the reliability in the estimated percentage
                for x in range(0, len(bestComboRanked) - 1):
                    if bestComboRanked[x][2] < statsAndProb[2]:  # Program scrolls through ranked combos and finds the appropriate spot to put this specified probability
                        bestComboRanked.insert(x, statsAndProb)
                        run = True
                        break
                if run == False:
                    bestComboRanked.append(statsAndProb)
                number = number + 1
            elif len(bestComboRanked) == 127:
                # print("works")
                run = False
                statsAndProb = [statOne, statTwo, currentProbStat, occurences]
                for x in range(0, len(bestComboRanked) - 1):
                    if bestComboRanked[x][2] < statsAndProb[
                        2]:  # once list is full, program scrolls through to find where some probabilites should be inserted
                        bestComboRanked.insert(x, statsAndProb)
                        bestComboRanked.pop(127)
                        break
            else:  # for first probability to be placed
                # print("testing")
                statsAndProb = [statOne, statTwo, currentProbStat, occurences]
                bestComboRanked.append(statsAndProb)

            if imRight > 0:  # An old thing I was doing before where I wanted the program to find the best combo and just return that
                if probStat < imRight / (imRight + wrong):
                    # if mixedValues < historicalMixedValues:
                    probStat = imRight / (imRight + wrong)
                    bestStat = [statOne, statTwo]
                    historicalMixedValues = (total) - (imRight + wrong)

            statTwo = statTwo + 1
        statOne = statOne + 1
    bestStatAndStatList = [probStat, bestStat]
    # print(historicalMixedValues)
    # return bestStatAndStatList
    return bestComboRanked


def avgLastXGames(gamesArray, team, numGames, date):  # Same As AvgSinceWhen method but a different method of doing it. Also allows for an easy way to calculate a teams current avg of the last X games
    if date == int(0):
        avgGame = ["", "", "", int(0), int(0), int(0), int(0), int(0), int(0), int(0), int(0), int(0), int(0), int(0),
                   int(0), int(0), int(0), int(0), int(0), int(0), int(0), int(0), int(0), int(0)]
        count = int(0)
        for x in range(0, 30):
            if gamesArray[x][0][0] == team:
                for y in range(152, 0, -1):
                    if gamesArray[x][y][0] != "":
                        count = count + 1
                        for z in range(5, 24):
                            avgGame[z] = avgGame[z] + gamesArray[x][y][z]
                    if count == numGames:
                        break
        for z in range(5, 24):
            avgGame[z] = avgGame[z] / numGames
    else:
        avgGame = ["", "", "", int(0), int(0), int(0), int(0), int(0), int(0), int(0), int(0), int(0), int(0), int(0),
                   int(0), int(0), int(0), int(0), int(0), int(0), int(0), int(0), int(0), int(0)]
        totalGames = int(0)
        for x in range(0, 30):
            if gamesArray[x][0][0] == team:
                for y in range(0, 154):
                    if gamesArray[x][y][2] == date and gamesArray[x][y + 1][0] != "":
                        z = int(y + 1)
                        while z < (y + numGames) and z < 153:
                            if gamesArray[x][z][0] != "":
                                totalGames = totalGames + 1
                                for a in range(5, 24):
                                    avgGame[a] = avgGame[a] + gamesArray[x][z][a]
                            if z < 153:
                                z = z + 1
                    elif gamesArray[x][y][2] == date and gamesArray[x][y + 1][0] == "":
                        totalGames = totalGames + 1
        if totalGames == 0:
            totalGames = 1
        for stat in range(5, 24):
            avgGame[stat] = round(avgGame[stat] / totalGames, 3)

    return avgGame


def whoWins(gamesArray, bestComboRevised, teamsList, numGames, date):  # Calculates and returns who wins
    winnerAndPercent = ["", ""]
    prob = int(0)
    avgTeam1 = avgLastXGames(gamesArray, teamsList[0], numGames, date)
    avgTeam2 = avgLastXGames(gamesArray, teamsList[1], numGames, date)
    for y in range(0, 136 - 9):
        if avgTeam1[bestComboRevised[y][0]] > avgTeam2[bestComboRevised[y][0]] and avgTeam1[bestComboRevised[y][1]] > avgTeam2[bestComboRevised[y][1]]:
            """print("WE FOUND A WINNNER")
            print(teamsList[0])"""
            winnerAndPercent[1] = bestComboRevised[y][2]
            winnerAndPercent[0] = teamsList[0]
            break
        elif avgTeam1[bestComboRevised[y][0]] < avgTeam2[bestComboRevised[y][0]] and avgTeam1[bestComboRevised[y][1]] < avgTeam2[bestComboRevised[y][1]]:
            """"print("WE FOUND A WINNNER")
            print(teamsList[1])
            print(bestComboRevised[x][2])"""
            winnerAndPercent[1] = bestComboRevised[y][2]
            winnerAndPercent[0] = teamsList[1]
            break
    if winnerAndPercent[0] == "":
        print("a probability could not be found. Roughly a 50/50 shot")  # in almost 5000 games this error is caught about 23 times. The error this causes is negligible
        winnerAndPercent[0] = teamsList[0]
    return winnerAndPercent


def testAccuracy(gamesArray, numGames, bestComboRevized):  # tests accuracy. Looks at the percentages I have calculated and backtracks to see how accurate this program was in predicting games
    teamsList = ["", ""]
    winnerAndPercent = ["", ""]
    totalGames = int(0)
    predictionCorrect = int(0)
    for x in range(0, 30):
        for y in range(0, 154):
            if gamesArray[x][y][0] != "":
                totalGames = totalGames + 1
                date = int(gamesArray[x][y][2])
                teamsList = [gamesArray[x][y][0], gamesArray[x][y][1][2]]
                winnerAndPercent = whoWins(gamesArray, bestComboRevised, teamsList, numGames, date)
                if winnerAndPercent[0] == gamesArray[x][y][0] and gamesArray[x][y][3] == 'W':
                    predictionCorrect = predictionCorrect + 1
                elif winnerAndPercent[0] == gamesArray[x][y][1][2] and gamesArray[x][y][3] == 'L':
                    predictionCorrect = predictionCorrect + 1
    return predictionCorrect / totalGames


# MAIN METHOD
workbook = load_workbook(filename="NBAGames2018to2020.xlsx")
print(workbook.sheetnames)
sheet = workbook.active
ptsAvgArray = []
gamesArray = [["" for x in range(154)] for y in range(30)]  # y represents amount of team. X represents the amount of games that teams have played. Its stored blank
a = int(0)
b = int(0)
blankGame = ["", "", "", int(0), int(0), int(0), int(0), int(0), int(0), int(0), int(0), int(0), int(0), int(0), int(0),
             int(0), int(0), int(0), int(0), int(0), int(0), int(0), int(0), int(0)]
for x in range(0, 30):
    for y in range(0, 154):
        gamesArray[x][y] = blankGame
for z in range(2, 4403): #imports games from excel file and sorts them into the games array
    gameList = ["", "", "", "", int(0), int(0), int(0), int(0), int(0), int(0), int(0), int(0), int(0), int(0), int(0),
                int(0), int(0), int(0), int(0), int(0), int(0), int(0), int(0), int(0)]
    for y in range(1, 23):
        cell = sheet.cell(row=z, column=y)
        x = y - 1
        if y == 2: #just formatting things to organize some of the data to make it more useable
            matchupArray = cell.value.split(' ')
            gameList[1] = matchupArray #seperates the Matchups into the home and away team
        elif y == 3:#turns the date into an integer so I can better sort the data in chronological order
            dateArray = str(cell.value).split('-')
            intDay = ''
            intDate = int(0)
            for x in range(0, 2):
                intDay = intDay + dateArray[2][x]
            dateArray[2] = intDay
            intDate = int(dateArray[0] + dateArray[1] + dateArray[2])
            gameList[2] = int(intDate)
        elif y > 5:
            gameList[x] = int(cell.value)
        else:
            gameList[x] = cell.value
    if a < 29:#function that sorts the games by team
        gamesArray[a][b] = gameList
        if b > 2:
            if gamesArray[a][b][0] != gamesArray[a][b - 1][0]:
                gamesArray[a][b] = ["", "", "", "", int(0), int(0), int(0), int(0), int(0), int(0), int(0), int(0),
                                    int(0), int(0), int(0), int(0), int(0), int(0), int(0), int(0), int(0), int(0),
                                    int(0), int(0)]
                a = a + 1
                b = 0
                gamesArray[a][b] = gameList

        b = b + 1
    else:
        for c in range(0, 30):
            if gamesArray[c][0][0] == gameList[0]:
                for d in range(0, 162):
                    if gamesArray[c][d][0] == "":
                        gamesArray[c][d] = gameList
                        break

for x in range(0, 30):
    print(gamesArray[x][0][0], end=" ")
print("")

gameStatOrder = ["Team", "matchup", "date", "win/loss", "mins", "points", "FGM", "FGA", "FG%", "3PM", "3PA", "3P%", "FTM", "FTA", "OREB",
                 "DREB", "TREB", "AST", "STL", "BLK", "TO", "PF"]
print(gameStatOrder)

numGames = int(input("What number of games do you want to analyze"))
print("these are the probabilites if you just used one stat to make your predictions")
print(gameStatOrder)
print(mostImportantStatRunAvg(gamesArray, numGames))
print("Now we will calculate the probabilites on multiple stats")
print("This could take a couple mins. Just running an analaysis of games based on your choice")
bestComboRevised = comboStatsRevized(gamesArray, gameStatOrder, numGames)

for x in range(0, len(bestComboRevised) - 1):
    print(bestComboRevised[x])
print("The notation are the numerical representations of the two stats, followed by the percent, followed by the amount of occurences this situation has occured")

print("The accuracy of using this combined method of predicting is", end=" ")
print(testAccuracy(gamesArray, numGames, bestComboRevised)) #spits out the percent of games you would get right if you followed this program

answer = "yes"

while answer != "no":
    teamsList = ["", ""]
    print("what probability do you want to know")
    print("Note: Enter teams in this format")

    teamsList[0] = input("Enter First Team")
    teamsList[1] = input("Enter Second Team")
    winnerAndPercent = whoWins(gamesArray, bestComboRevised, teamsList, numGames, 0)
    print("The winner and percent chance of winning is", end=" ")
    print(winnerAndPercent[0], end=" ")
    print(winnerAndPercent[1])
    print("say 'no' if you do not want to continue")
    answer = input("Do you want to go again")
